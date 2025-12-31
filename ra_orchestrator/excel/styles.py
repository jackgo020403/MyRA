"""Excel styling utilities."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Color palette
COLOR_TITLE = "1F4E78"  # Dark blue
COLOR_MEMO_BG = "E7E6E6"  # Light gray
COLOR_MEMO_CONCLUSION = "C00000"  # Red for conclusion text
COLOR_HEADER_ROW = "4472C4"  # Blue for ledger headers
COLOR_Q1 = "FFF2CC"  # Light yellow
COLOR_Q2 = "E2EFDA"  # Light green
COLOR_Q3 = "FCE4D6"  # Light orange
COLOR_Q4 = "DDEBF7"  # Light blue
COLOR_Q5 = "F4B084"  # Light peach

QUESTION_COLORS = {
    "Q1": COLOR_Q1,
    "Q2": COLOR_Q2,
    "Q3": COLOR_Q3,
    "Q4": COLOR_Q4,
    "Q5": COLOR_Q5,
}


def get_question_color(q_id: str) -> str:
    """Get color for a question ID."""
    return QUESTION_COLORS.get(q_id, "FFFFFF")  # White default


def apply_title_style(ws, row: int, col_start: int, col_end: int):
    """Apply title styling to merged cell range."""
    cell = ws.cell(row, col_start)
    cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge cells
    ws.merge_cells(
        start_row=row,
        start_column=col_start,
        end_row=row,
        end_column=col_end
    )


def apply_memo_style(ws, start_row: int, end_row: int, col_start: int, col_end: int):
    """Apply memo block styling."""
    for row in range(start_row, end_row + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill(start_color=COLOR_MEMO_BG, end_color=COLOR_MEMO_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Bold headers
    ws.cell(start_row, col_start).font = Font(name="Calibri", size=11, bold=True)


def apply_conclusion_style(cell):
    """Apply conclusion text style (red, bold)."""
    cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_MEMO_CONCLUSION)


def apply_decomposition_style(ws, start_row: int, end_row: int, col_start: int, col_end: int, q_id: str):
    """Apply question decomposition styling with background color."""
    color = get_question_color(q_id)

    for row in range(start_row, end_row + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Bold Q ID
    ws.cell(start_row, col_start).font = Font(name="Calibri", size=11, bold=True)


def apply_header_row_style(ws, row: int, col_start: int, col_end: int):
    """Apply ledger header row styling."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER_ROW, end_color=COLOR_HEADER_ROW, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border


def apply_ledger_row_style(ws, row: int, col_start: int, col_end: int, q_id: str, row_type: str):
    """Apply styling to a ledger data row."""
    color = get_question_color(q_id)

    for col in range(col_start, col_end + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # If HEADER row type, make bold
    if row_type == "HEADER":
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", size=10, bold=True)


def set_column_widths(ws, num_columns: int):
    """Set appropriate column widths."""
    # Default widths for meta columns
    column_widths = {
        1: 8,   # Row_ID
        2: 12,  # Row_Type
        3: 10,  # Question_ID
        4: 15,  # Section
        5: 50,  # Statement
        6: 15,  # Supports_Row_IDs
        7: 30,  # Source_URL
        8: 20,  # Source_Name
        9: 12,  # Date
        10: 12, # Confidence
        11: 20, # Notes
    }

    # Dynamic columns get default width
    for col_idx in range(1, num_columns + 1):
        width = column_widths.get(col_idx, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
