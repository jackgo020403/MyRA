"""Excel writer for research output."""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from ra_orchestrator.state import RAState, ResearchPlan, LedgerSchema, MemoBlock, LedgerRow
from ra_orchestrator.excel.styles import (
    apply_title_style,
    apply_memo_style,
    apply_conclusion_style,
    apply_decomposition_style,
    apply_header_row_style,
    apply_ledger_row_style,
    set_column_widths,
)


def write_dry_run_excel(state: RAState, output_dir: Path) -> str:
    """
    Write a dry-run Excel file with Title, Memo placeholder, Decomposition, and empty ledger.

    Args:
        state: RA state with research_plan and ledger_schema
        output_dir: Directory to save Excel file

    Returns:
        Path to saved Excel file
    """
    plan: ResearchPlan = state["research_plan"]
    schema: LedgerSchema = state["ledger_schema"]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Output"

    current_row = 1

    # ========== TITLE ==========
    num_cols = len(schema.meta_columns) + len(schema.dynamic_columns)
    ws.cell(current_row, 1, plan.research_title)
    apply_title_style(ws, current_row, 1, num_cols)
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Blank row
    current_row += 1

    # ========== MEMO BLOCK (PLACEHOLDER) ==========
    memo_start_row = current_row

    ws.cell(current_row, 1, "EXECUTIVE MEMO")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    ws.cell(current_row, 1, "Key Conclusion:")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    ws.cell(current_row, 1, "[PLACEHOLDER - Will be auto-generated after research]")
    apply_conclusion_style(ws.cell(current_row, 1))
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
    current_row += 1

    current_row += 1
    ws.cell(current_row, 1, "Key Supporting Evidence:")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    for i in range(1, 4):
        ws.cell(current_row, 1, f"  {i}. [Evidence summary] (Source: [Publisher, Year], Row ID: [#])")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

    current_row += 1
    ws.cell(current_row, 1, "Caveat/Confidence:")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    ws.cell(current_row, 1, "[Optional - if needed]")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
    current_row += 1

    memo_end_row = current_row - 1
    apply_memo_style(ws, memo_start_row, memo_end_row, 1, num_cols)

    # Blank row
    current_row += 1

    # ========== QUESTION DECOMPOSITION ==========
    ws.cell(current_row, 1, "QUESTION DECOMPOSITION")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    for sub_q in plan.sub_questions:
        decomp_start = current_row

        ws.cell(current_row, 1, f"{sub_q.q_id}: {sub_q.question}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        ws.cell(current_row, 1, f"  Rationale: {sub_q.rationale}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        decomp_end = current_row - 1
        apply_decomposition_style(ws, decomp_start, decomp_end, 1, num_cols, sub_q.q_id)

        # Small gap between questions
        current_row += 1

    # Blank row
    current_row += 1

    # ========== LEDGER ==========
    ws.cell(current_row, 1, "RESEARCH LEDGER")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    # Header row
    ledger_header_row = current_row
    all_columns = schema.meta_columns + [col.name for col in schema.dynamic_columns]

    for col_idx, col_name in enumerate(all_columns, start=1):
        ws.cell(current_row, col_idx, col_name)

    apply_header_row_style(ws, current_row, 1, len(all_columns))
    current_row += 1

    # Placeholder rows
    ws.cell(current_row, 1, "[Ledger rows will be populated during research phase]")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
    current_row += 1

    # ========== FREEZE PANES ==========
    # Freeze at the ledger header row so Title + Memo + Decomposition stay visible
    ws.freeze_panes = ws.cell(ledger_header_row + 1, 1)

    # ========== COLUMN WIDTHS ==========
    set_column_widths(ws, num_cols)

    # ========== SAVE ==========
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"research_output_dryrun_{timestamp}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)

    return str(filepath)


def write_full_excel(state: RAState, output_dir: Path) -> str:
    """
    Write the full Excel file with actual research data (Milestone 2).

    Args:
        state: Complete RA state with ledger_rows
        output_dir: Directory to save Excel file

    Returns:
        Path to saved Excel file
    """
    plan: ResearchPlan = state["research_plan"]
    schema: LedgerSchema = state["ledger_schema"]
    ledger_rows = state.get("ledger_rows", [])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Output"

    current_row = 1
    num_cols = len(schema.meta_columns) + len(schema.dynamic_columns)

    # ========== TITLE ==========
    ws.cell(current_row, 1, plan.research_title)
    apply_title_style(ws, current_row, 1, num_cols)
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    # ========== MEMO BLOCK (Placeholder for M3) ==========
    memo_start_row = current_row
    ws.cell(current_row, 1, "EXECUTIVE MEMO")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    ws.cell(current_row, 1, "Key Conclusion:")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    ws.cell(current_row, 1, "[Auto-generated in Milestone 3 - Synthesizer]")
    apply_conclusion_style(ws.cell(current_row, 1))
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
    current_row += 2

    ws.cell(current_row, 1, "Key Supporting Evidence:")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    for i in range(1, 4):
        ws.cell(current_row, 1, f"  {i}. [Evidence extraction in M3]")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

    apply_memo_style(ws, memo_start_row, current_row - 1, 1, num_cols)
    current_row += 1

    # ========== QUESTION DECOMPOSITION ==========
    ws.cell(current_row, 1, "QUESTION DECOMPOSITION")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    for sub_q in plan.sub_questions:
        decomp_start = current_row
        ws.cell(current_row, 1, f"{sub_q.q_id}: {sub_q.question}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        ws.cell(current_row, 1, f"  Rationale: {sub_q.rationale}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        apply_decomposition_style(ws, decomp_start, current_row - 1, 1, num_cols, sub_q.q_id)
        current_row += 1

    current_row += 1

    # ========== LEDGER HEADER ==========
    ws.cell(current_row, 1, "RESEARCH LEDGER")
    ws.cell(current_row, 1).font = Font(name="Calibri", size=12, bold=True)
    current_row += 1

    ledger_header_row = current_row
    all_columns = schema.meta_columns + [col.name for col in schema.dynamic_columns]

    for col_idx, col_name in enumerate(all_columns, start=1):
        ws.cell(current_row, col_idx, col_name)

    apply_header_row_style(ws, current_row, 1, len(all_columns))
    current_row += 1

    # ========== LEDGER DATA ROWS ==========
    if ledger_rows:
        print(f"\n[EXCEL] Writing {len(ledger_rows)} ledger rows...")
        for row_data in ledger_rows:
            ws.cell(current_row, 1, row_data.row_id)
            ws.cell(current_row, 2, row_data.row_type)
            ws.cell(current_row, 3, row_data.question_id)
            ws.cell(current_row, 4, row_data.section)
            ws.cell(current_row, 5, row_data.statement)
            ws.cell(current_row, 6, row_data.supports_row_ids or "")
            ws.cell(current_row, 7, row_data.source_url or "")
            ws.cell(current_row, 8, row_data.source_name or "")
            ws.cell(current_row, 9, row_data.date or "")
            ws.cell(current_row, 10, row_data.confidence or "")
            ws.cell(current_row, 11, row_data.notes or "")

            # Dynamic columns
            for col_idx, col in enumerate(schema.dynamic_columns, start=12):
                value = row_data.dynamic_fields.get(col.name, "")
                ws.cell(current_row, col_idx, value)

            # Apply styling
            apply_ledger_row_style(
                ws, current_row, 1, len(all_columns),
                row_data.question_id, row_data.row_type
            )

            current_row += 1

    # ========== FREEZE PANES & FORMATTING ==========
    ws.freeze_panes = ws.cell(ledger_header_row + 1, 1)
    set_column_widths(ws, len(all_columns))

    # ========== SAVE ==========
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"research_output_{timestamp}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)
    print(f"[EXCEL] Saved to: {filepath}")

    return str(filepath)
