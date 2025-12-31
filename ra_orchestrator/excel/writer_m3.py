"""
Excel writer for Milestone 3 - Adds SYNTHESIS and MEMO tabs.
"""

import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ra_orchestrator.state import RAState, QuestionSynthesis, MemoBlock


def write_milestone3_excel(state: RAState, base_excel_path: str, output_dir: Path) -> str:
    """
    Add SYNTHESIS and MEMO tabs to existing Excel file.

    Args:
        state: RA state with question_syntheses and memo_block
        base_excel_path: Path to existing Excel with EVIDENCE tab
        output_dir: Output directory

    Returns:
        Path to new Excel file with synthesis and memo
    """
    syntheses = state.get("question_syntheses", [])
    memo = state.get("memo_block")
    ledger_rows = state.get("ledger_rows", [])

    # Load existing workbook
    wb = load_workbook(base_excel_path)

    # Add SYNTHESIS tab
    if "SYNTHESIS" in wb.sheetnames:
        del wb["SYNTHESIS"]
    ws_synthesis = wb.create_sheet("SYNTHESIS", 0)  # Insert as first sheet
    _write_synthesis_tab(ws_synthesis, syntheses, ledger_rows)

    # Add MEMO tab
    if "MEMO" in wb.sheetnames:
        del wb["MEMO"]
    ws_memo = wb.create_sheet("MEMO", 0)  # Insert before SYNTHESIS
    _write_memo_tab(ws_memo, memo, state["research_plan"].research_title)

    # Transform "Research Output" sheet to "RAW DATA"
    if "Research Output" in wb.sheetnames:
        ws_raw = wb["Research Output"]
        ws_raw.title = "RAW DATA"
        _transform_to_raw_data(ws_raw, ledger_rows)

    # Save as new file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"research_final_m3_{timestamp}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)
    return str(filepath)


def _write_synthesis_tab(ws, syntheses: list, ledger_rows: list):
    """Write SYNTHESIS tab with per-question analysis.

    Args:
        ws: Worksheet
        syntheses: List of QuestionSynthesis
        ledger_rows: List of LedgerRow for creating hyperlinks
    """

    # Create evidence lookup map (row_id -> LedgerRow)
    evidence_map = {row.row_id: row for row in ledger_rows}

    # Title
    ws.cell(1, 1, "SYNTHESIS - Per-Question Analysis")
    ws.cell(1, 1).font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30

    current_row = 3

    for synthesis in syntheses:
        # Question ID and Question
        ws.cell(current_row, 1, f"{synthesis.question_id}: {synthesis.question}")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        ws.cell(current_row, 1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

        # Mini Conclusion
        ws.cell(current_row, 1, "Mini Conclusion:")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(current_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1

        ws.cell(current_row, 1, synthesis.mini_conclusion)
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.row_dimensions[current_row].height = 60
        current_row += 1

        # Logical Reasoning
        ws.cell(current_row, 1, "Logical Reasoning:")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(current_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1

        for reasoning in synthesis.logical_reasoning:
            # Parse citation: "(Source: SourceName, Evidence #15)"
            # Extract evidence ID to create hyperlink
            match = re.search(r'Evidence #(\d+)', reasoning)

            if match:
                evidence_id = int(match.group(1))
                evidence_row = evidence_map.get(evidence_id)

                if evidence_row:
                    # Replace citation with hyperlinked version
                    # Format: "Statement text (Source Name)"
                    citation_pattern = r'\(Source: ([^,]+), Evidence #\d+\)'
                    replacement = lambda m: f'({m.group(1)})'
                    text_with_source = re.sub(citation_pattern, replacement, reasoning)

                    # Create cell with text
                    cell = ws.cell(current_row, 1, text_with_source)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                    # Add hyperlink to source name portion
                    # Find where the source name is in the text
                    source_match = re.search(r'\(([^)]+)\)$', text_with_source)
                    if source_match and evidence_row.source_url:
                        cell.hyperlink = evidence_row.source_url
                        cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
                else:
                    # Evidence not found, just display as-is
                    cell = ws.cell(current_row, 1, reasoning)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                # No citation found, display as-is
                cell = ws.cell(current_row, 1, reasoning)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws.row_dimensions[current_row].height = max(30, len(reasoning) // 100 * 15)
            current_row += 1

        # Supporting Evidence IDs
        ws.cell(current_row, 1, "Supporting Evidence:")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(current_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1

        evidence_ids = ", ".join([f"#{id}" for id in synthesis.supporting_evidence_ids])
        ws.cell(current_row, 1, f"Evidence rows: {evidence_ids}")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 1

        # Confidence
        ws.cell(current_row, 1, "Confidence:")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(current_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1

        confidence_color = {
            "High": "C6E0B4",
            "Medium": "FFE699",
            "Low": "F4B084"
        }.get(synthesis.confidence, "FFFFFF")

        ws.cell(current_row, 1, f"{synthesis.confidence}: {synthesis.confidence_rationale}")
        ws.cell(current_row, 1).fill = PatternFill(start_color=confidence_color, end_color=confidence_color, fill_type="solid")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

        # Spacing
        current_row += 2

    # Column widths
    ws.column_dimensions['A'].width = 120
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def _write_memo_tab(ws, memo: MemoBlock, research_title: str):
    """Write MEMO tab with executive summary."""

    # Title
    ws.cell(1, 1, research_title)
    ws.cell(1, 1).font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 40

    ws.cell(2, 1, "EXECUTIVE SUMMARY")
    ws.cell(2, 1).font = Font(name="Calibri", size=14, bold=True)
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30

    current_row = 4

    # Executive Summary
    _write_section(ws, current_row, "Executive Summary", [memo.executive_summary])
    current_row += 4

    # Key Findings
    _write_section(ws, current_row, "Key Findings", memo.key_findings)
    current_row += len(memo.key_findings) + 3

    # Cross-Question Insights
    _write_section(ws, current_row, "Cross-Question Insights", memo.cross_question_insights)
    current_row += len(memo.cross_question_insights) + 3

    # Implications
    _write_section(ws, current_row, "Implications", memo.implications)
    current_row += len(memo.implications) + 3

    # Methodology Note
    _write_section(ws, current_row, "Methodology & Limitations", [memo.methodology_note])

    # Column widths
    ws.column_dimensions['A'].width = 120
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def _write_section(ws, start_row: int, title: str, items: list):
    """Write a section with title and bullet points."""
    # Section title
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(start_row, 1).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    ws.cell(start_row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{start_row}:F{start_row}")
    ws.row_dimensions[start_row].height = 25

    # Items
    for i, item in enumerate(items, 1):
        row = start_row + i
        ws.cell(row, 1, f"• {item}")
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = max(30, len(item) // 100 * 15)


def _transform_to_raw_data(ws, ledger_rows: list):
    """
    Transform existing Research Output sheet to RAW DATA format.

    Changes:
    1. Remove title/memo/decomposition sections (keep only data)
    2. Reorganize by source in MLA format
    3. Update title to "RAW DATA"

    Args:
        ws: Worksheet to transform
        ledger_rows: List of LedgerRow objects
    """
    # Clear entire sheet
    ws.delete_rows(1, ws.max_row)

    # Write new title
    ws.cell(1, 1, "RAW DATA")
    ws.cell(1, 1).font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 35

    current_row = 3

    # Group evidence by source URL
    from collections import defaultdict
    by_source = defaultdict(list)

    for row in ledger_rows:
        source_key = (row.source_url, row.source_name, row.date)
        by_source[source_key].append(row)

    # Sort sources alphabetically by source name
    sorted_sources = sorted(by_source.items(), key=lambda x: x[0][1])  # Sort by source_name

    # Write each source in MLA format
    for (source_url, source_name, date), evidence_list in sorted_sources:
        # MLA Citation Header (clickable hyperlink)
        citation = f"{source_name} ({date})"

        cell = ws.cell(current_row, 1, citation)
        cell.font = Font(name="Calibri", size=12, bold=True, color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add hyperlink
        if source_url:
            cell.hyperlink = source_url

        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Evidence items from this source
        for ev in evidence_list:
            # Format: "• [Statement] (Evidence #ID, Q: question_id)"
            statement_text = f"• {ev.statement} (Evidence #{ev.row_id}, Q: {ev.question_id})"

            cell = ws.cell(current_row, 1, statement_text)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws.row_dimensions[current_row].height = max(30, len(statement_text) // 100 * 15)
            current_row += 1

        # Spacing between sources
        current_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 120
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
